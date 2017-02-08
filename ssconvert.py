import subprocess
import hashlib
import logging
import sys
import os.path
import xlrd, csv 
from openpyxl import load_workbook
#_dirname = os.path.dirname(os.path.dirname(os.path.realpath(sys.argv[0])))
#sys.path.insert(0, os.path.join(_dirname, "lib"))
#LOGDIR  = os.path.join(_dirname, "logs")
#LOGFILE = os.path.join(LOGDIR, "excel.log")
logger = logging.getLogger()
handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s %(name)-12s %(levelname)-8s %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)
logger.setLevel(logging.DEBUG)


class CONVERT(object):
    def  __init__(self):
        self.ssconvert_bin_path = "/usr/bin/ssconvert"
  
    def xl_to_csv(self, inpfile, outfile):
        """
        It uses Gnumeric's ssconvert utility to convert xls/xlsx files to csv.
        Expected location of ssconvert binary :- /usr/bin/ssconvert
        Usage :- xl_to_csv('Example.xls')
        It will generate Example.xls.csv as converted csv file.
        Return code values :-
        0 = Success.
        1 = Failure.
        """
        #output_filename="{0}.csv".format('.'.join(filename.split('.')[:-1]))
        proc=subprocess.Popen([self.ssconvert_bin_path, inpfile, outfile],
                              stdin=subprocess.PIPE,
                              stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE,
                              )
        proc.communicate()
        return proc.returncode

    def get_xlsheet_info(self, inpfile):
        xltype = inpfile.split('.')[-1]
        if xltype == 'xls':
            workbook = xlrd.open_workbook(inpfile)
            sheet_names = workbook.sheet_names()
            xl_sheet = workbook.sheet_by_name(sheet_names[0]) 
            rows = xl_sheet.nrows
            cols = xl_sheet.ncols
            return rows, cols, xl_sheet, 'xls'
        elif xltype == 'xlsx':
            wb = load_workbook(inpfile)
            sheet = wb.worksheets[0]
            row_count = sheet.max_row
            column_count = sheet.max_column
            return row_count, column_count, sheet, 'xlsx'
        return 0, 0, '', ''

    def read_csv_row_cols(self, outfile):
        row_count, col_count = 0, 0
        with open(outfile, "r") as f:
            reader = csv.reader(f, delimiter = ",")
            data = list(reader)
            row_count = len(data)
            if data:
                col_count = len(data[0])
        return row_count, col_count


    def get_hash(self, inp_ls):
        ### call clean up functions like strip '', '
        #inp_ls = [ele.strip('"') for ele in inp_ls]
        inp_str = ','.join(inp_ls)
        hash_object = hashlib.md5(inp_str)
        return hash_object.hexdigest()

    def compare_contents_xlsx(self, xlobj, outfile):
        xl_rows = list(xlobj.iter_rows())
        flg = 1
        with open(outfile, "r") as f:
            reader = csv.reader(f, delimiter = ",")
            for i, csv_line in enumerate(reader):
                row_index = xl_rows[i]
                xl_line = [str(ele.value) for ele in row_index]
                hash1 = self.get_hash(csv_line)
                hash2 = self.get_hash(xl_line)
                if hash1 != hash2:
                    flg = 0
                    break
        if flg == 1:
            return True
        return False 

    def compare_contents_xls(self, xl_cols, xlobj, outfile):
        flg = 1
        with open(outfile, "r") as f:
            reader = csv.reader(f, delimiter = ",")
            for i, csv_line in enumerate(reader):
                xl_line = [xlobj.cell(i, j).value for j in range(xl_cols)]
                hash1 = self.get_hash(csv_line)
                hash2 = self.get_hash(xl_line)
                if hash1 != hash2:
                    flg = 0
                    break
        if flg == 1:
            return True
        return False 

    def match_lines(self, x, y, msg):
        if int(x) == int(y):
            logger.info("{0} count matched".format(msg))
            return True
        else:
            logger.error("{0} count not matching".format(msg))
            return False

    def process_conversion(self, inpfile, outfile):
        if not os.path.isfile(self.ssconvert_bin_path):
            logger.error("{0} does not exist.".format(self.ssconvert_bin_path))
            return False 

        if not os.path.exists(inpfile):
            logger.error("{0} file does not exists.".format(inpfile))
            return False
        try:
            status_code = self.xl_to_csv(inpfile, outfile)
        except Exception as e:
            logger.error("Error in conversion {0}".format(e))
            return False

        if status_code != 0:
            logger.error("failed conversion")
            return False
        logger.info("conversion success csv path {0}".format(outfile))

        xl_rows, xl_cols, xlobj, xltype = self.get_xlsheet_info(inpfile)
        csv_row, csv_col = self.read_csv_row_cols(outfile)
        logger.info("EXCEL ROWS = {0} COLS = {1}".format(xl_rows, xl_cols))
        logger.info("CSV ROWS = {0} COLS = {1}".format(csv_row, csv_col))
        match_flg_row = self.match_lines(xl_rows, csv_row, 'Row')
        match_flg_col = self.match_lines(xl_cols, csv_col, 'Column')
        if (not match_flg_row) or (not match_flg_col): return False
        if xltype == 'xlsx' and self.compare_contents_xlsx(xlobj, outfile):
            logger.info("contents matched")
            return True
        elif xltype == 'xls' and self.compare_contents_xls(xl_cols, xlobj, outfile):
            logger.info("contents matched")
            return True
        logger.error("contents didnt match")
        return False
     

if __name__ == '__main__':
    inpfile = sys.argv[1]
    outfile = sys.argv[2]
    obj = CONVERT()
    obj.process_conversion(inpfile, outfile)
